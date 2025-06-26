# Title: This script include classes for the review page of the HWTT Analysis Tool.
#
# Author: Farhad Abdollahi (farhad.abdollahi.ctr@dot.gov)
# Date: 06/24/2025
# ======================================================================================================================

# Importing the required libraries.
import os
import sys
import shutil
import sqlite3
import fnmatch
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from PyQt5.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QAction, QDoubleSpinBox, QLabel, \
    QPushButton, QWidget, QGridLayout, QFormLayout, QLineEdit, QFileDialog, QMessageBox, QGroupBox, QProgressBar, \
    QPlainTextEdit, QStackedWidget, QCheckBox, QComboBox, QTabWidget, QScrollArea, QTableWidget, QTableWidgetItem, \
    QSpinBox, QFrame, QDialog, QRadioButton, QButtonGroup, QInputDialog
from PyQt5.QtGui import QPixmap, QFont, QRegExpValidator, QDoubleValidator, QIntValidator, QPixmap
from PyQt5.QtCore import Qt, QRegExp
from qtwidgets import AnimatedToggle
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from scripts.Alg01_UtilityFunctions import Binary_to_Array, Read_Resize_Image
from scripts.Alg02_SQL_Manager import Get_DB_SummaryData, Get_Identifier_Combinations
from scripts.Alg03_HWTT_Analysis_Functions import HWTT_Analysis, ModelPower, TsengLyttonModel, YinModel


class DB_ReviewPage(QMainWindow):
    """
    This class design a review database page in which the database is shown as table where user can select any specific 
    sample and revise its analysis procedure. 
    """
    def __init__(self, conn, cursor, DB_Name, DB_Folder, stack, shared_data):
        # Initiate the required parameters.
        super().__init__()
        self.conn = conn            # connection to the SQL database.
        self.cursor = cursor        # cursor for running the SQL commands. 
        self.DB_Name = DB_Name      # Name of the database.
        self.DB_Folder = DB_Folder  # Directory at which the database is saved. 
        self.stack = stack
        self.shared_data = shared_data
        self.stack.currentChanged.connect(self.Function_Button_Sync_Summary_Info)
        self.ColumnNames = ['B-number', 'Lane #', 'Lift Location', 'Lab Aging', 'Rep #', 'Wheel Side',
                            'Max Rutting (mm)', 'Max Passes', 
                            'Rutting @ 10k (mm) [2PP]', 'Rutting @ 20k (mm) [2PP]', 
                            'Stripping Number [2PP]', 'SIP [2PP]', 'SIP @ 12.5 mm [2PP]', 
                            'Creep Slope [2PP]', 'Stripping Slope [2PP]', 
                            'Rutting @ 10k (mm) [Yin]', 'Rutting @ 20k (mm) [Yin]', 
                            'Stripping Number [Yin]', 'SIP [Yin]', 'SIP @ 12.5 mm [Yin]', 
                            'Creep Slope [Yin]', 'Stripping Slope [Yin]', 
                            'Rutting @ 10k (mm) [6deg]', 'Rutting @ 20k (mm) [6deg]', 
                            'Stripping Number [6deg]', 'Creep Slope [6deg]', 'Is Outlier?', 'ID']
        self.SQL_ColumnNames = [
            'Bnumber', 'Lane_Num', 'Lift_Location', 'Lab_Aging', 'RepNumber', 'Wheel_Side',
            'TPP_Max_Rut_mm', 'TPP_Max_Pass', 
            'TPP_RuttingAt10k_mm', 'TPP_RuttingAt20k_mm', 
            'TPP_StrippingNumber', 'TPP_SIP', 'TPP_SIP_Adj', 
            'TPP_CreepLine_Slope', 'TPP_StrippingLine_Slope', 
            'Yin_RuttingAt10k_mm', 'Yin_RuttingAt20k_mm', 
            'Yin_StrippingNumber', 'Yin_SIP', 'Yin_SIP_Adj', 
            'Yin_CreepLine_Slope', 'Yin_StrippingLine_Slope', 
            'Poly6_RuttingAt10k_mm', 'Poly6_RuttingAt20k_mm', 
            'Poly6_StrippingNumber', 'Poly6_CreepLine_Slope', 'IsOutlier', 'id']
        # self.ColumnNamesAnalysis = [
        #     'B_Number', 'Lab_Aging_Condition', 'Num_Data', 
        #     'ICO_Baseline_mean', 'ICO_Baseline_std', 'ICO_Baseline_COV', 
        #     'ICO_Deconv_mean', 'ICO_Deconv_std', 'ICO_Deconv_COV', 
        #     'ICO_Tangential_mean', 'ICO_Tangential_std', 'ICO_Tangential_COV', 
        #     'ISO_Baseline_mean', 'ISO_Baseline_std', 'ISO_Baseline_COV', 
        #     'ISO_Deconv_mean', 'ISO_Deconv_std', 'ISO_Deconv_COV', 
        #     'ISO_Tangential_mean', 'ISO_Tangential_std', 'ISO_Tangential_COV', 
        #     'Aliphatic_Area_Baseline_mean', 'Aliphatic_Area_Baseline_std', 'Aliphatic_Area_Baseline_COV', 
        #     'Aliphatic_Area_Tangential_mean', 'Aliphatic_Area_Tangential_std', 'Aliphatic_Area_Tangential_COV', 
        #     'Carbonyl_Peak_Wavenumber_mean', 'Carbonyl_Peak_Wavenumber_std', 'Carbonyl_Peak_Wavenumber_COV', 
        #     'Sulfoxide_Peak_Wavenumber_mean', 'Sulfoxide_Peak_Wavenumber_std', 'Sulfoxide_Peak_Wavenumber_COV', 
        #     'Carbonyl_Peak_Absorption_mean', 'Carbonyl_Peak_Absorption_std', 'Carbonyl_Peak_Absorption_COV', 
        #     'Sulfoxide_Peak_Absorption_mean', 'Sulfoxide_Peak_Absorption_std', 'Sulfoxide_Peak_Absorption_COV']
        self.IdentifierCombs = Get_Identifier_Combinations(self.cursor)
        self.initUI()
    # ------------------------------------------------------------------------------------------------------------------
    def initUI(self):
        # Initiate the user interface. 
        # Main widget and layout
        main_widget = QWidget()
        layout = QHBoxLayout(main_widget)                   # Make a horizontal layout.
        self.setCentralWidget(main_widget)
        # self.setStyleSheet("background-color: #f0f0f0;")
        # Generate the left and right layouts, where left one include the main table of results, and right one include 
        #   the sections for: Summary of DB, Filter and find, and action buttons. 
        # --------------------------------------------------------------------------------------------------------------
        # ---------------- Left Layout (Main table) --------------------------------------------------------------------
        # --------------------------------------------------------------------------------------------------------------
        # Section 01 (Left Layout): Main Table.
        SectT01 = QGroupBox("Review Table")
        SectT01.setStyleSheet("QGroupBox { font-weight: bold; }")
        SectT01_Layout = QVBoxLayout()
        self.SectT01_Label_NumRecords = QLabel('Number of records shown in table: 0')
        self.SectT01_Label_NumRecords.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.SectT01_Label_CurrentSelection = QLabel("Current selection: N/A")
        self.SectT01_Label_CurrentSelection.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        # Define the table. 
        self.Table = QTableWidget()
        self.Table.setRowCount(10)
        self.Table.setColumnCount(len(self.ColumnNames))
        self.Table.setHorizontalHeaderLabels(self.ColumnNames)
        self.Table.setSelectionBehavior(self.Table.SelectRows)
        self.Table.setSelectionMode(self.Table.SingleSelection)
        self.Table.selectionModel().selectionChanged.connect(self.Function_Update_CurrentSelection)
        # Place them inside the section. 
        SectT01_Layout.addWidget(self.SectT01_Label_NumRecords)
        SectT01_Layout.addWidget(self.SectT01_Label_CurrentSelection)
        SectT01_Layout.addWidget(self.Table)
        SectT01.setLayout(SectT01_Layout)
        layout.addWidget(SectT01, 75)
        # --------------------------------------------------------------------------------------------------------------
        # ------------- Right Layout (Sections for Summary of DB, Filter and Search, and Action Buttons) ---------------
        # --------------------------------------------------------------------------------------------------------------
        SummaryData = Get_DB_SummaryData(self.cursor)
        Right_Layout = QVBoxLayout()
        # Section 02 (Right Layout, Top section): Summary of DB.
        SectT02 = QGroupBox("Summary of DB")
        SectT02.setStyleSheet("QGroupBox { font-weight: bold; }")
        SectT02_Layout = QVBoxLayout()
        SummayText_Label = QLabel("This section shows a summary of the analyzed data available in the Database.")
        SummayText_Label.setWordWrap(True)
        SectT02_Layout.addWidget(SummayText_Label)
        SectT02_FormLayout = QFormLayout()          # Define a form layout.
        # Create the labels in Section 02.
        Label01 = QLabel("Number of data:")
        self.Label_NumData = QLabel(f'{SummaryData["NumRows"]}')
        Label02 = QLabel("Number of valid data:")
        self.Label_NumValidData = QLabel(f'{SummaryData["NumValidRows"]}')
        Label03 = QLabel("Avg Number Replicates:")
        self.Label_AvgNumReplicates = QLabel(f'{SummaryData["AvgNumRep"]:.1f}')
        Label04 = QLabel("Number of unique B-numbers:")
        self.Label_NumUniqueBnum = QLabel(f'{SummaryData["NumUniqueBnumber"]}')
        Label05 = QLabel("Number of unique Lab agings:")
        self.Label_NumUniqueLabAge = QLabel(f'{SummaryData["NumUniqueLabAging"]}')
        Label06 = QLabel("Number of Unique Bnum/LabAge:")
        self.Label_NumUniqueBnumLabAge = QLabel(f'{SummaryData["NumUniqueBnumLabAge"]}')
        # Syncing button.
        self.Button_Sync = QPushButton("Refresh summay")
        self.Button_Sync.setFont(QFont("Arial", 8))
        self.Button_Sync.clicked.connect(self.Function_Button_Sync_Summary_Info)
        self.Button_Sync.setFixedSize(100, 30)
        self.Button_Sync.setStyleSheet(
        """
        QPushButton:hover {background-color: lightgray;}
        QPushButton:pressed {background-color: gray;}
        """)
        # Place the labels in the GUI.
        SectT02_FormLayout.addRow(Label01,  self.Label_NumData)
        SectT02_FormLayout.addRow(Label02,  self.Label_NumValidData)
        SectT02_FormLayout.addRow(Label03,  self.Label_AvgNumReplicates)
        SectT02_FormLayout.addRow(Label04, self.Label_NumUniqueBnum)
        SectT02_FormLayout.addRow(Label05, self.Label_NumUniqueLabAge)
        SectT02_FormLayout.addRow(Label06, self.Label_NumUniqueBnumLabAge)
        SectT02_Layout.addLayout(SectT02_FormLayout)
        SectT02_Layout.addWidget(self.Button_Sync, alignment=Qt.AlignHCenter | Qt.AlignTop)
        SectT02.setLayout(SectT02_Layout)
        Right_Layout.addWidget(SectT02, 25)
        # --------------------------------------------------------------------------------------------------------------
        # --------------------------------------------------------------------------------------------------------------
        # Section 03 (Right Layout, Middle section): Search and Filter the DB. 
        SectT03 = QGroupBox("Search and Filter")
        SectT03.setStyleSheet("QGroupBox { font-weight: bold; }")
        SectT03_Layout = QVBoxLayout()
        # First, dropdown for the B-numbers. 
        Label_DropDown01 = QLabel("Select the asphalt binder B-number:")
        self.DropDown_Bnumber = QComboBox()
        self.DropDown_Bnumber.addItems(['Please select...', 'All mixtures'] + 
                                       list(self.IdentifierCombs['Bnumber'].unique()))
        self.DropDown_Bnumber.activated.connect(self.Function_DropDown_Bnumber)
        # Second dropdown for the Lane number.
        Label_DropDown02 = QLabel("Select the lane number:")
        self.DropDown_LaneNumber = QComboBox()
        self.DropDown_LaneNumber.addItems(
            ['All lanes'] + [f'Lane {ii}' for ii in sorted(list(self.IdentifierCombs['Lane_Num'].unique()))])
        self.DropDown_LaneNumber.setEnabled(True)
        self.DropDown_LaneNumber.activated.connect(self.Function_DropDown_LaneNumber) 
        # Next, dropdown for the Laboratory aging state. 
        Label_DropDown03 = QLabel("Select the laboratory aging level:")
        self.DropDown_LabAging = QComboBox()
        self.DropDown_LabAging.addItems(['All aging levels'])
        self.DropDown_LabAging.setEnabled(False)
        self.DropDown_LabAging.activated.connect(self.Function_DropDown_LabAging)
        # Finally, add the apply button. 
        self.Button_Fetch = QPushButton("Fetch data")
        self.Button_Fetch.setStyleSheet(
        """
        QPushButton:hover {background-color: lightgray;}
        QPushButton:pressed {background-color: gray;}
        """)
        self.Button_Fetch.clicked.connect(self.Function_Button_Fetch)
        self.Button_Fetch.setFixedWidth(150)
        # Place the items in the window. 
        SectT03_Layout.addWidget(Label_DropDown01)
        SectT03_Layout.addWidget(self.DropDown_Bnumber)
        SectT03_Layout.addWidget(Label_DropDown02)
        SectT03_Layout.addWidget(self.DropDown_LaneNumber)
        SectT03_Layout.addWidget(Label_DropDown03)
        SectT03_Layout.addWidget(self.DropDown_LabAging)
        SectT03_Layout.addWidget(self.Button_Fetch, alignment=Qt.AlignHCenter)
        SectT03.setLayout(SectT03_Layout)
        Right_Layout.addWidget(SectT03, 25)
        # --------------------------------------------------------------------------------------------------------------
        # --------------------------------------------------------------------------------------------------------------
        # Section 04 (Right Layout, Lower section): Action buttons for Delete and Modify the record. 
        SectT04 = QGroupBox("Action Buttons")
        SectT04.setStyleSheet("QGroupBox { font-weight: bold; }")
        SectT04_Layout = QVBoxLayout()
        # Button for Deleting a record.
        self.Button_Delete = QPushButton("Delete Record")
        self.Button_Delete.setFont(QFont("Arial", 11, QFont.Bold))
        self.Button_Delete.clicked.connect(self.Function_Delete_Record)
        self.Button_Delete.setFixedSize(150, 30)
        self.Button_Delete.setEnabled(True)
        self.Button_Delete.setStyleSheet(
        """
        QPushButton:enabled {background-color: #FA8072; color: black;}
        QPushButton:hover {background-color: lightgray;}
        QPushButton:pressed {background-color: gray;}
        QPushButton:checked {background-color: lime;}
        """)
        # Button for Modify a record.
        self.Button_Modify = QPushButton("Modify Record")
        self.Button_Modify.setFont(QFont("Arial", 11, QFont.Bold))
        self.Button_Modify.clicked.connect(self.Function_Modify_Record)
        self.Button_Modify.setFixedSize(150, 30)
        self.Button_Modify.setEnabled(True)
        self.Button_Modify.setStyleSheet(
        """
        QPushButton:enabled {background-color: #1E90FF; color: black;}
        QPushButton:hover {background-color: lightgray;}
        QPushButton:pressed {background-color: gray;}
        QPushButton:checked {background-color: lime;}
        """)
        SectT04_Layout.addWidget(self.Button_Delete, alignment=Qt.AlignHCenter | Qt.AlignTop)
        SectT04_Layout.addWidget(self.Button_Modify, alignment=Qt.AlignHCenter | Qt.AlignTop)
        SectT04.setLayout(SectT04_Layout)
        Right_Layout.addWidget(SectT04, 15)
        # --------------------------------------------------------------------------------------------------------------
        # --------------------------------------------------------------------------------------------------------------
        # Section 05 (Right Layout, Lower section): Buttons to move to the other pages.
        SectT05 = QGroupBox("Navigator to Other Pages")
        SectT05.setStyleSheet("QGroupBox { font-weight: bold; }")
        SectT05_Layout = QVBoxLayout()
        # Button for Return to the main page.
        self.Button_Main = QPushButton("Main Page")
        self.Button_Main.setFont(QFont("Arial", 11, QFont.Bold))
        self.Button_Main.clicked.connect(self.Function_Button_MainPage)
        self.Button_Main.setFixedSize(160, 35)
        self.Button_Main.setEnabled(True)
        self.Button_Main.setStyleSheet(
        """
        QPushButton:enabled {background-color: #FFE4B5; color: black;}
        QPushButton:hover {background-color: lightgray;}
        QPushButton:pressed {background-color: gray;}
        QPushButton:checked {background-color: lime;}
        """)
        # Button for move to the analysis page.
        self.Button_Analysis = QPushButton("Analysis Page")
        self.Button_Analysis.setFont(QFont("Arial", 11, QFont.Bold))
        self.Button_Analysis.clicked.connect(self.Function_Button_MainPage)
        self.Button_Analysis.setFixedSize(160, 35)
        self.Button_Analysis.setEnabled(True)
        self.Button_Analysis.setStyleSheet(
        """
        QPushButton:enabled {background-color: #FFE4B5; color: black;}
        QPushButton:hover {background-color: lightgray;}
        QPushButton:pressed {background-color: gray;}
        QPushButton:checked {background-color: lime;}
        """)
        SectT05_Layout.addWidget(self.Button_Main, alignment=Qt.AlignHCenter | Qt.AlignTop)
        SectT05_Layout.addWidget(self.Button_Analysis, alignment=Qt.AlignHCenter | Qt.AlignTop)
        SectT05.setLayout(SectT05_Layout)
        Right_Layout.addWidget(SectT05, 15)
        # --------------------------------------------------------------------------------------------------------------
        # --------------------------------------------------------------------------------------------------------------
        # Section 06 (Right Layout, Lower section): Export section.
        SectT06 = QGroupBox("Export Section")
        SectT06.setStyleSheet("QGroupBox { font-weight: bold; }")
        SectT06_Layout = QVBoxLayout()
        # Button for export the individual record.
        self.Button_ExportOne = QPushButton("Export (Individual Record)")
        self.Button_ExportOne.setFont(QFont("Arial", 11, QFont.Bold))
        self.Button_ExportOne.clicked.connect(self.Function_Button_Export_Individual)
        self.Button_ExportOne.setFixedSize(220, 35)
        self.Button_ExportOne.setEnabled(True)
        self.Button_ExportOne.setStyleSheet(
        """
        QPushButton:hover {background-color: lightgray;}
        QPushButton:pressed {background-color: gray;}
        QPushButton:checked {background-color: lime;}
        """)
        # Button for export the individual record.
        self.Button_ExportDB = QPushButton("Export (Database)")
        self.Button_ExportDB.setFont(QFont("Arial", 11, QFont.Bold))
        self.Button_ExportDB.clicked.connect(self.Function_Button_Export_Database)
        self.Button_ExportDB.setFixedSize(220, 35)
        self.Button_ExportDB.setEnabled(True)
        self.Button_ExportDB.setStyleSheet(
        """
        QPushButton:hover {background-color: lightgray;}
        QPushButton:pressed {background-color: gray;}
        QPushButton:checked {background-color: lime;}
        """)
        self.Export_ProgressBar = QProgressBar()
        self.Export_ProgressBar.setMinimum(0)
        self.Export_ProgressBar.setMaximum(0)
        self.Export_ProgressBar.hide()
        SectT06_Layout.addWidget(self.Button_ExportOne, alignment=Qt.AlignHCenter | Qt.AlignTop)
        SectT06_Layout.addWidget(self.Button_ExportDB,  alignment=Qt.AlignHCenter | Qt.AlignTop)
        SectT06_Layout.addWidget(self.Export_ProgressBar)
        SectT06.setLayout(SectT06_Layout)
        Right_Layout.addWidget(SectT06, 20)
        layout.addLayout(Right_Layout, 25)
    # ==================================================================================================================
    # =========================== Define the Functions =================================================================
    # ==================================================================================================================
    def Function_Button_Sync_Summary_Info(self):
        # Get the latest summary information. 
        SummaryData = Get_DB_SummaryData(self.cursor)
        # Update the values. 
        self.Label_NumData.setText(f'{SummaryData["NumRows"]}')
        self.Label_NumValidData.setText(f'{SummaryData["NumValidRows"]}')
        self.Label_AvgNumReplicates.setText(f'{SummaryData["AvgNumRep"]:.1f}')
        self.Label_NumUniqueBnum.setText(f'{SummaryData["NumUniqueBnumber"]}')
        self.Label_NumUniqueLabAge.setText(f'{SummaryData["NumUniqueLabAging"]}')
        self.Label_NumUniqueBnumLabAge.setText(f'{SummaryData["NumUniqueBnumLabAge"]}')
        # ----------------------------
        # Also update the identifiers. 
        self.IdentifierCombs = Get_Identifier_Combinations(self.cursor)         # Update the Identifiers.
        # Update the dropdown menu as well. 
        self.DropDown_Bnumber.clear()           # Clear and reset the B-number dropdown menu. 
        self.DropDown_Bnumber.addItems(['Please Select...', 'All mixtures'] + 
                                       list(self.IdentifierCombs['Bnumber'].unique()))
        self.DropDown_Bnumber.setCurrentIndex(0)
        self.DropDown_Bnumber.setEnabled(True)
        self.DropDown_LaneNumber.clear()        # Clear and reset the lane number dropdown menu. 
        self.DropDown_LaneNumber.addItems(
            ['All Lanes'] + [f'Lane {ii}' for ii in sorted(list(self.IdentifierCombs['Lane_Num'].unique()))])
        self.DropDown_LaneNumber.setCurrentIndex(0)
        self.DropDown_LaneNumber.setEnabled(True)
        self.DropDown_LabAging.clear()
        self.DropDown_LabAging.addItems(['All Aging Levels'])
        self.DropDown_LabAging.setCurrentIndex(0)
        self.DropDown_LabAging.setEnabled(False)
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Button_Fetch(self):
        """
        This function reads the filters and then update the table with the information from the DB. 
        """
        # First clear the table and row selector. 
        self.Function_Clear_Tables()
        self.Table.clearSelection()
        self.SectT01_Label_CurrentSelection.setText("Current selection: N/A")
        # Read the filters. 
        Bnumber    = self.DropDown_Bnumber.currentText()
        LaneNumber = self.DropDown_LaneNumber.currentText()
        LabAging   = self.DropDown_LabAging.currentText()
        if self.DropDown_Bnumber.currentIndex() == 0:
            # When B-number dropdown is "Please select...", the table should remain empty. 
            return
        elif self.DropDown_Bnumber.currentIndex() == 1:
            # When B-number dropdown is "All mixtures", the table should be filled accordingly, considering other 
            #   dropdown values.
            if self.DropDown_LaneNumber.currentIndex() == 0:
                # It means all lane numbers. 
                if self.DropDown_LabAging.currentIndex() == 0:          # Considering all lab aging levels. 
                    self.cursor.execute(f"SELECT {', '.join(self.SQL_ColumnNames)} FROM HWTT")
                elif self.DropDown_LabAging.currentIndex() > 0:         # Considering a specific lab aging level. 
                    self.cursor.execute(f"SELECT {', '.join(self.SQL_ColumnNames)} FROM HWTT WHERE Lab_Aging = ?", 
                               (LabAging,))
                else:
                    raise ValueError(f'Unexpected error occured! ' + 
                                     f'Lab_Aging index={self.DropDown_LabAging.currentIndex()}')
            elif self.DropDown_LaneNumber.currentIndex() > 0:
                # It means a specific lane is selected. 
                LaneNo = int(LaneNumber[5:])
                if self.DropDown_LabAging.currentIndex() == 0:          # Considering all lab aging levels. 
                    self.cursor.execute(f"SELECT {', '.join(self.SQL_ColumnNames)} FROM HWTT WHERE Lane_Num = ?", 
                               (LaneNo,))
                elif self.DropDown_LabAging.currentIndex() > 0:         # Considering a specific lab aging level. 
                    self.cursor.execute(
                        f"SELECT {', '.join(self.SQL_ColumnNames)} FROM HWTT WHERE Lab_Aging = ?, Lane_Num = ?", 
                        (LabAging, LaneNo))
                else:
                    raise ValueError(f'Unexpected error occured! ' + 
                                     f'Lab_Aging index={self.DropDown_LabAging.currentIndex()}')
            else:
                raise ValueError(f'Unexpected error occured! ' + 
                                 f'Lane_Num index={self.DropDown_LaneNumber.currentIndex()}')
        elif self.DropDown_Bnumber.currentIndex() > 1:
            # When a specific B-number is selected, the table should be filled accordingly, considering other dropdowns, 
            #   it is noted in this case, we don't have to worry about the lane numbers, as the B-number is specific to 
            #   a certain lane number.
            if self.DropDown_LabAging.currentIndex() == 0:      # For all aging levels. 
                self.cursor.execute(f"SELECT {', '.join(self.SQL_ColumnNames)} FROM HWTT WHERE Bnumber = ?", (Bnumber,))
            elif self.DropDown_LabAging.currentIndex() > 0:     # For a specific lab aging level. 
                self.cursor.execute(
                    f"SELECT {', '.join(self.SQL_ColumnNames)} FROM HWTT WHERE Bnumber = ?, Lab_Aging = ?", 
                    (Bnumber, LabAging))
            else:
                raise ValueError(f'Unexpected error occured! ' + 
                                 f'Lab_Aging index={self.DropDown_LabAging.currentIndex()}')
        else:
            raise ValueError(f'Unexpected error occured! ' + 
                             f'Bnumber index={self.DropDown_Bnumber.currentIndex()}')
        # Get the rows. 
        Rows = self.cursor.fetchall()
        # Modify the table with the row value
        self.SectT01_Label_NumRecords.setText(f'Number of records shown in table: {len(Rows)}')
        self.Table.setRowCount(len(Rows))
        for row_idx, row_data in enumerate(Rows):
            for col_idx, cell_data in enumerate(row_data):
                if type(cell_data) == float:
                    item = QTableWidgetItem(f'{cell_data:.4f}')
                else:
                    item = QTableWidgetItem(str(cell_data))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.Table.setItem(row_idx, col_idx, item)
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Clear_Tables(self):
        """
        This function clears the main table. 
        """
        self.Table.clearSelection()         # Clear the row selection in table. 
        self.SectT01_Label_NumRecords.setText('Number of fetched data (rows): 0')
        self.Table.setRowCount(10)          # Set the number of rows to 10 empty rows. 
        for row_idx in range(10):
            for col_idx in range(len(self.ColumnNames)):
                self.Table.setItem(row_idx, col_idx, QTableWidgetItem(''))
    # ------------------------------------------------------------------------------------------------------------------
    def Function_DropDown_Bnumber(self):
        """
        This function called when the user change another B-number. 
        """
        # First, clear the main table.
        self.Function_Clear_Tables()         
        # Depending on the selection, update the other dropdowns. 
        if self.DropDown_Bnumber.currentIndex() == 0:
            # "Please select..." is selected. 
            self.DropDown_LaneNumber.setEnabled(True)
            self.DropDown_LaneNumber.setCurrentIndex(0)
            self.DropDown_LabAging.clear()
            self.DropDown_LabAging.addItems(['All aging levels'])
            self.DropDown_LabAging.setEnabled(False)
            self.DropDown_LabAging.setCurrentIndex(0)
        elif self.DropDown_Bnumber.currentIndex() == 1:
            # "All mixtures" is selected. 
            self.DropDown_Bnumber.clear()
            self.DropDown_Bnumber.addItems(['Please select...', 'All mixtures'] + 
                                           sorted(list(self.IdentifierCombs['Bnumber'].unique())))
            self.DropDown_Bnumber.setCurrentIndex(1)
            self.DropDown_Bnumber.setEnabled(True)
            self.DropDown_LaneNumber.clear()
            self.DropDown_LaneNumber.addItems(
                ['All lanes'] + [f'Lane {ii}' for ii in sorted(list(self.IdentifierCombs['Lane_Num'].unique()))])
            self.DropDown_LaneNumber.setCurrentIndex(0)
            self.DropDown_LaneNumber.setEnabled(True)
            self.DropDown_LabAging.clear()
            self.DropDown_LabAging.addItems(['All aging levels'] + 
                                            sorted(list(self.IdentifierCombs['Lab_Aging'].unique())))
            self.DropDown_LabAging.setEnabled(True)
            self.DropDown_LabAging.setCurrentIndex(0)
        elif self.DropDown_Bnumber.currentIndex() > 1:
            # A specific B-number is selected. 
            Bnumber = self.DropDown_Bnumber.currentText()
            TempDF = self.IdentifierCombs[self.IdentifierCombs['Bnumber'] == Bnumber]
            LaneNo = TempDF.iloc[0, 1]
            self.DropDown_LaneNumber.clear()
            self.DropDown_LaneNumber.addItems(['All lanes', f'Lane {LaneNo}'])
            self.DropDown_LaneNumber.setCurrentIndex(1)
            self.DropDown_LaneNumber.setEnabled(True)
            self.DropDown_LabAging.clear()
            self.DropDown_LabAging.addItems(['All aging levels'] + sorted(list(TempDF['Lab_Aging'].unique())))
            self.DropDown_LabAging.setCurrentIndex(0)
            self.DropDown_LabAging.setEnabled(True)
        else:
            pass
    # ------------------------------------------------------------------------------------------------------------------
    def Function_DropDown_LaneNumber(self):
        """
        This function called when the user select another lane number, change selection. 
        """
        # First, clear the main table.
        self.Function_Clear_Tables()
        # Depending on the selection, update the other dropdowns. 
        if self.DropDown_LaneNumber.currentIndex() == 0:
            # "All lanes" is selected. 
            self.DropDown_Bnumber.clear()
            self.DropDown_Bnumber.addItems(['Please select...', 'All mixtures'] + 
                                           list(self.IdentifierCombs['Bnumber'].unique()))
            self.DropDown_Bnumber.setCurrentIndex(0)
            self.DropDown_Bnumber.setEnabled(True)
            self.DropDown_LaneNumber.clear()
            self.DropDown_LaneNumber.addItems(
                ['All lanes'] + [f'Lane {ii}' for ii in sorted(list(self.IdentifierCombs['Lane_Num'].unique()))])
            self.DropDown_LaneNumber.setCurrentIndex(0)
            self.DropDown_LaneNumber.setEnabled(True)
            self.DropDown_LabAging.clear()
            self.DropDown_LabAging.addItems(['All aging levels'])
            self.DropDown_LabAging.setCurrentIndex(0)
            self.DropDown_LabAging.setEnabled(False)
        elif self.DropDown_LaneNumber.currentIndex() > 0:
            # A specific lane number is selected. 
            LaneNo = int(self.DropDown_LaneNumber.currentText()[5:])
            TempDF = self.IdentifierCombs[self.IdentifierCombs['Lane_Num'] == LaneNo]
            self.DropDown_Bnumber.clear()
            self.DropDown_Bnumber.addItems(['Please select...', 'All mixtures'] + list(TempDF['Bnumber'].unique()))
            self.DropDown_Bnumber.setCurrentIndex(1)
            self.DropDown_Bnumber.setEnabled(True)
            self.DropDown_LaneNumber.clear()
            self.DropDown_LaneNumber.addItems(['All lanes', f'Lane {LaneNo}'])
            self.DropDown_LaneNumber.setCurrentIndex(1)
            self.DropDown_LaneNumber.setEnabled(True)
            self.DropDown_LabAging.clear()
            self.DropDown_LabAging.addItems(['All aging levels'] + sorted(list(TempDF['Lab_Aging'].unique())))
            self.DropDown_LabAging.setCurrentIndex(0)
            self.DropDown_LabAging.setEnabled(True)
        else:
            pass
    # ------------------------------------------------------------------------------------------------------------------
    def Function_DropDown_LabAging(self):
        """
        This function called when the user select another lab aging level, where it needs to clear the table. 
        """
        # Only clear the table. 
        self.Function_Clear_Tables()
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Button_MainPage(self):
        """
        This function changes the stack view from the Review page to the Main page. 
        """
        self.stack.setCurrentIndex(0)  # Switch to the first page, Main page.
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Update_CurrentSelection(self, selected, deselected):
        """
        This function will update the current selection label. 
        """
        try:
            # Get the index of the selected row. 
            for index in selected.indexes():
                idx = index.row()
                break
            # Update the label. 
            if self.Table.item(idx, 1).text() == '':
                self.SectT01_Label_CurrentSelection.setText("Current selection: N/A") 
            else:
                self.SectT01_Label_CurrentSelection.setText(
                    f"Current selection: B-number={self.Table.item(idx, 0).text()} " +
                    f"for Lane {self.Table.item(idx, 1).text()} ({self.Table.item(idx, 2).text()} lift), " + 
                    f'at age level of "{self.Table.item(idx, 3).text()}", Rep {self.Table.item(idx, 4).text()}')
        except:
            pass
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Delete_Record(self):
        """
        This function deletes the current selection from the database. 
        """
        # Find the selected index. 
        SelectedIndices = self.Table.selectionModel().selectedIndexes()
        if len(SelectedIndices) == 0:           
            # Nothing is selected. 
            QMessageBox.critical(self, "Data Selection Error!", 
                                        f"Row was not selected. Please first select the row you want to delete " + 
                                        f"from the database.")
            return
        idx = SelectedIndices[0].row()
        # Check the id value. 
        ID = self.Table.item(idx, self.Table.columnCount() - 1)
        if ID == None:
            # Table is empty. 
            QMessageBox.critical(self, "Data Selection Error!", 
                                        f"Selected row ({idx + 1}) is empty. Please first fetch the data using the " +
                                        f'"Search and Filter" section, then select the intended row, and then click ' +
                                        f'"Delete Record" button.')
            return
        else:
            ID = int(ID.text())
            Msg  = f'Do you want to Permanently Delete the following record?:\n' + \
                   f'B-number={self.Table.item(idx, 0).text()} ' + \
                   f"for Lane {self.Table.item(idx, 1).text()} ({self.Table.item(idx, 2).text()} lift), " + \
                   f'at age level of "{self.Table.item(idx, 3).text()}", Rep {self.Table.item(idx, 4).text()}'
            Question = QMessageBox()
            Question.setIcon(QMessageBox.Question)
            Question.setWindowTitle("Delete Record Confirmation")
            Question.setText(Msg)
            Question.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            Question.setDefaultButton(QMessageBox.No)
            Reply = Question.exec_()
            # Check the response. 
            if Reply == QMessageBox.Yes:
                # User Choose Yes.
                # deleting the record from the Database.
                self.cursor.execute("DELETE FROM HWTT WHERE id = ?", (ID,))
                self.conn.commit()
                # Updating the table. 
                self.Function_Button_Fetch()
            else:
                # User clicked No, and ignore the deletion. 
                print("Ignore")
                return
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Modify_Record(self):
        pass
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Button_Export_Individual(self):
        """
        This function performs the exporting the results into a Excel spreadsheet.  
        """
        self.Export_ProgressBar.show()
        # Find the selected index. 
        SelectedIndices = self.Table.selectionModel().selectedIndexes()
        if len(SelectedIndices) == 0:           
            # Nothing is selected. 
            QMessageBox.critical(self, "Data Selection Error!", 
                                        f"Row was not selected. Please first select the row you want to export " + 
                                        f"from the database.")
            self.Export_ProgressBar.hide()
            return
        idx = SelectedIndices[0].row()
        # Check the id value. 
        ID = self.Table.item(idx, self.Table.columnCount() - 1)
        if ID == None or ID.text() == '':
            # Table is empty. 
            QMessageBox.critical(self, "Data Selection Error!", 
                                        f"Selected row ({idx + 1}) is empty. Please first fetch the data using the " +
                                        f'"Search and Filter" section, then select the intended row, and then click ' +
                                        f'"Export (Individual Record)" button.')
            self.Export_ProgressBar.hide()
            return
        # Otherwise, everything is ready for exporting. 
        ID = int(ID.text())             # The ID to retrieve the data from the DB. 
        # -------------------------------------------------------
        # Part 1: Ask for a place to save the file and file name. 
        Directory = QFileDialog.getExistingDirectory(self, "Please select Saving Directory", "")
        # If a file is selected by the user, update the Input_SavePath.
        if not Directory:
            QMessageBox.critical(self, "Directory Selection Failed!", f"Directory was NOT selected. Please try again.")
            self.Export_ProgressBar.hide()
            return
        print(f'Saving Directory: {Directory}')
        # # Freeze the main window. 
        # self.setEnabled(False)
        # Ask for the file name. 
        self.cursor.execute("SELECT FileName FROM HWTT WHERE id = ?", (ID,))

        FileName, IsOkButtonPressed = QInputDialog.getText(self,
                                                           "Output File Name", 
                                                           "Please enter the output file name (without .xlsx):",
                                                           text=os.path.splitext(self.cursor.fetchone()[0])[0])
        if IsOkButtonPressed:
            FileName = FileName + '.xlsx'
            print(f"Saving File Name: {FileName}")
        else:
            QMessageBox.critical(self, "Output File Name Failed!", 
                                 f"Output file name was NOT confirmed. Please try again.")
            self.Export_ProgressBar.hide()
            return
        # --------------------------------------------------------------------------------------------------------------
        # Prepare the output file. 
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        # Set the title of the sheet
        ws.title = "Sheet1"
        # Define some styles. 
        Info_fill       = PatternFill(start_color="FFE989", end_color="FFCC00", fill_type="solid")
        TPP_fill        = PatternFill(start_color="A7E2FF", end_color="A7E2FF", fill_type="solid")
        Yin_fill        = PatternFill(start_color="F4FEBA", end_color="F4FEBA", fill_type="solid")
        Poly_fill       = PatternFill(start_color="FDBBBB", end_color="FDBBBB", fill_type="solid")
        Data_Fill       = PatternFill(start_color="DFDED9", end_color="DFDED9", fill_type="solid")
        Invalid_Fill    = PatternFill(start_color="FF1515", end_color="FF1515", fill_type="solid")
        thin            = Side(border_style="thin", color="000000")
        cell_border     = Border(top=thin, left=thin, right=thin, bottom=thin)
        header_font     = Font(name="Arial", bold=True, size=11, color="000000")
        cell_font       = Font(name="Arial", size=11, color="000000")
        center_alignment= Alignment(horizontal="center", vertical="center")
        left_alignment  = Alignment(horizontal="left", vertical="center")
        # ----------------------------------------------------------------------------
        # Extract the General information parameters. 
        ColNames_Info = [
            'Bnumber', 'Lane_Num', 'Lift_Location', 'Lab_Aging', 'RepNumber', 'Wheel_Side', 'FileName', 'FileDirectory', 
            'Test_Name', 'Technician_Name', 'Test_Date', 'Test_Time', 'Test_Condition', 
            'Target_Test_Temperature', 'Avg_Test_Temperature', 'Std_Test_Temperature', 
            'Valid_Min_Pass', 'Valid_Max_Pass', 'Other_Comments', 'IsOutlier']
        Labels_Info   = [
            'B-number', 'Lane number', 'Lift location', 'Lab aging level', 'Repetition number', 
            'Testing wheel side', 'Raw data file name', 'Raw data directory', 
            'Test name (by operator)', 'Technician name', 'Testing date', 'Testing time', 'Testing condition (wet/dry)', 
            'Target testing temperature', 'Avg. of testing temperature', 'Std. of testing temperature', 
            'Index of first valid pass', 'Index of last valid pass', 
            'Any other comments/notes', 'Is this test considered Outlier']
        self.cursor.execute(f'SELECT {", ".join(ColNames_Info)} FROM HWTT WHERE id = ?', (ID,))
        Values_Info = list(self.cursor.fetchone())
        # Add precision to the testing temperatures. 
        Values_Info[13] = f'{Values_Info[13]:.2f} °C'
        Values_Info[14] = f'{Values_Info[14]:.2f} °C'
        Values_Info[15] = f'{Values_Info[15]:.2f} °C'
        # Make isOutlier yes/no.
        Values_Info[19] = 'Yes' if Values_Info[19] else "No"
        # ------------------------------------------------------
        # Write the General information. 
        ws.merge_cells('A1:B1')
        cell = ws.cell(row=1, column=1, value='General Information')
        cell.fill = Info_fill
        cell.border = cell_border
        cell.font = Font(name="Arial", size=13, bold=True, color="000000")
        cell.alignment = center_alignment
        for i, (header, value) in enumerate(zip(Labels_Info, Values_Info), start=2):
            # Write the title.
            cell1 = ws.cell(row=i, column=1, value=header + ':')
            cell1.fill = Info_fill
            cell1.border = cell_border
            cell1.font = header_font
            cell1.alignment = left_alignment
            # Write the value. 
            cell1 = ws.cell(row=i, column=2, value=value)
            cell1.fill = Info_fill
            cell1.border = cell_border
            cell1.font = cell_font
            cell1.alignment = left_alignment
        NextRowIndex = i + 2
        # ----------------------------------------------------------------------------------------------------------------------
        # Go to the TPP model results. 
        # Extract the results. 
        ColNames_2PP = [
            'TPP_StrippingNumber', 'TPP_Max_Rut_mm', 'TPP_Max_Pass', 
            'TPP_RuttingAt10k_mm', 'TPP_RuttingAt20k_mm', 
            'TPP_ModelCoeff_a', 'TPP_ModelCoeff_b', 
            'TPP_ModelCoeff_alpha', 'TPP_ModelCoeff_beta', 'TPP_ModelCoeff_gamma', 'TPP_ModelCoeff_Phi', 
            'TPP_Stripping_Rutting_mm', 'TPP_SIP', 'TPP_SIP_Yvalue', 'TPP_SIP_Adj', 'TPP_SIP_Adj_Yvalue', 
            'TPP_CreepLine_Slope', 'TPP_CreepLine_Intercept', 'TPP_StrippingLine_Slope', 'TPP_StrippingLine_Intercept', 
            'TPP_StrippingLine_Slope_Adj', 'TPP_StrippingLine_Intercept_Adj']
        Labels_2PP = [
            'Stripping number (SN)', 'Maximum rutting (mm)', 'Maximum passes', 
            'Rutting @ 10k passes (mm)', 'Rutting @ 20k passes (mm)', 
            'Model coefficient, a', 'Model coefficient, b', 
            'Model coefficient, α', 'Model coefficient, β', 'Model coefficient, γ', 'Model coefficient, Φ', 
            'Stripping rutting (mm)', 'SIP', 'Rutting @ SIP (mm)', 'Adjusted SIP (@ 12.5 mm rutting)', 'Rutting @ adjusted SIP', 
            'Creep line slope', 'Creep line intercept (mm)', 'Stripping line slope', 'Stripping line intercept (mm)', 
            'Adjusted stripping line slope', 'Adjusted stripping line intercept (mm)'
        ]
        self.cursor.execute(f'SELECT {", ".join(ColNames_2PP)} FROM HWTT WHERE id = ?', (ID,))
        Values_2PP = list(self.cursor.fetchone())
        # ------------------------------------------------------
        # Write the General information. 
        ws.merge_cells(f'A{NextRowIndex}:B{NextRowIndex}')
        cell = ws.cell(row=NextRowIndex, column=1, value='Two-Part Power (2PP) Model')
        cell.fill = TPP_fill
        cell.border = cell_border
        cell.font = Font(name="Arial", size=13, bold=True, color="000000")
        cell.alignment = center_alignment
        for i in range(1, 5):
            for j in range(2):
                cell1 = ws.cell(row=NextRowIndex+i, column=j+1, value='')
                cell1.fill = TPP_fill
        # Put the image. 
        Image_Obj = Read_Resize_Image('./assets/Two-Part Power (2PP) Equation.png', 72)
        ws.add_image(Image_Obj, f"A{NextRowIndex+1}")
        NextRowIndex += 5
        # Write the 2PP model parameters. 
        for i, (header, value) in enumerate(zip(Labels_2PP, Values_2PP), start=NextRowIndex):
            # Write the title.
            cell1 = ws.cell(row=i, column=1, value=header + ':')
            cell1.fill = TPP_fill
            cell1.border = cell_border
            cell1.font = header_font
            cell1.alignment = left_alignment
            # Write the value. 
            cell1 = ws.cell(row=i, column=2, value=value)
            cell1.fill = TPP_fill
            cell1.border = cell_border
            cell1.font = cell_font
            cell1.alignment = left_alignment
        NextRowIndex = i + 2
        # ----------------------------------------------------------------------------------------------------------------------
        # Go to the Yin et al. model results. 
        # Extract the results. 
        ColNames_Yin = [
            'Yin_StrippingNumber', 'Yin_RuttingAt10k_mm', 'Yin_RuttingAt20k_mm', 
            'Yin_Parameter_LCSN', 'Yin_Parameter_LCST', 'Yin_Parameter_DeltaEpsAt10k',
            'Yin_ModelCoeff_Step1_ro', 'Yin_ModelCoeff_Step1_LCult', 'Yin_ModelCoeff_Step1_beta', 
            'Yin_ModelCoeff_Step2_RutMax', 'Yin_ModelCoeff_Step2_alpha', 'Yin_ModelCoeff_Step2_lambda', 
            'Yin_ModelCoeff_Step3_Eps0', 'Yin_ModelCoeff_Step3_theta', 
            'Yin_Stripping_Rutting_mm', 'Yin_SIP', 'Yin_SIP_Yvalue', 'Yin_SIP_Adj', 'Yin_SIP_Adj_Yvalue', 
            'Yin_CreepLine_Slope', 'Yin_CreepLine_Intercept', 'Yin_StrippingLine_Slope', 'Yin_StrippingLine_Intercept', 
            'Yin_StrippingLine_Slope_Adj', 'Yin_StrippingLine_Intercept_Adj']
        Labels_Yin = [
            'Stripping number (SN)', 'Rutting @ 10k passes (mm)', 'Rutting @ 20k passes (mm)', 
            'Analysis Parameter, LCSN', 'Analysis Parameter, LCST', 'Analysis Parameter, Δεp @ 10k',
            'Model coefficient, ρ', 'Model coefficient, LCult', 'Model coefficient, β', 
            'Model coefficient, Rut∞', 'Model coefficient, α', 'Model coefficient, λ', 
            'Model coefficient, ε0', 'Model coefficient, θ', 
            'Stripping rutting (mm)', 'SIP', 'Rutting @ SIP (mm)', 'Adjusted SIP (@ 12.5 mm rutting)', 'Rutting @ adjusted SIP', 
            'Creep line slope', 'Creep line intercept (mm)', 'Stripping line slope', 'Stripping line intercept (mm)', 
            'Adjusted stripping line slope', 'Adjusted stripping line intercept (mm)']
        self.cursor.execute(f'SELECT {", ".join(ColNames_Yin)} FROM HWTT WHERE id = ?', (ID,))
        Values_Yin = list(self.cursor.fetchone())
        # ------------------------------------------------------
        # Write the General information. 
        ws.merge_cells(f'A{NextRowIndex}:B{NextRowIndex}')
        cell = ws.cell(row=NextRowIndex, column=1, value='Yin et al. (2014) Model')
        cell.fill = Yin_fill
        cell.border = cell_border
        cell.font = Font(name="Arial", size=13, bold=True, color="000000")
        cell.alignment = center_alignment
        for i in range(1, 10):
            for j in range(2):
                cell1 = ws.cell(row=NextRowIndex+i, column=j+1, value='')
                cell1.fill = Yin_fill
        # Put the image. 
        Image_Obj = Read_Resize_Image('./assets/Yin et al Equation.png', 153)
        ws.add_image(Image_Obj, f"A{NextRowIndex+1}")
        NextRowIndex += 9
        # Write the 2PP model parameters. 
        for i, (header, value) in enumerate(zip(Labels_Yin, Values_Yin), start=NextRowIndex):
            # Write the title.
            cell1 = ws.cell(row=i, column=1, value=header + ':')
            cell1.fill = Yin_fill
            cell1.border = cell_border
            cell1.font = header_font
            cell1.alignment = left_alignment
            # Write the value. 
            cell1 = ws.cell(row=i, column=2, value=value)
            cell1.fill = Yin_fill
            cell1.border = cell_border
            cell1.font = cell_font
            cell1.alignment = left_alignment
        NextRowIndex = i + 2
        # ----------------------------------------------------------------------------------------------------------------------
        # Go to the 6th degree polynomial model results. 
        # Extract the results. 
        ColNames_Poly = [
            'Poly6_StrippingNumber', 'Poly6_RuttingAt10k_mm', 'Poly6_RuttingAt20k_mm', 
            'Poly6_ModelCoeff_a0', 'Poly6_ModelCoeff_a1', 'Poly6_ModelCoeff_a2', 'Poly6_ModelCoeff_a3', 
            'Poly6_ModelCoeff_a4', 'Poly6_ModelCoeff_a5', 'Poly6_ModelCoeff_a6', 
            'Poly6_Stripping_Rutting_mm', 'Poly6_CreepLine_Slope', 'Poly6_CreepLine_Intercept']
        Labels_Poly = [
            'Stripping number (SN)', 'Rutting @ 10k passes (mm)', 'Rutting @ 20k passes (mm)', 
            'Model coefficient, a0', 'Model coefficient, a1', 'Model coefficient, a2', 
            'Model coefficient, a3', 'Model coefficient, a4', 'Model coefficient, a5', 'Model coefficient, a6',
            'Stripping rutting (mm)', 'Creep line slope', 'Creep line intercept (mm)']
        self.cursor.execute(f'SELECT {", ".join(ColNames_Poly)} FROM HWTT WHERE id = ?', (ID,))
        Values_Poly = list(self.cursor.fetchone())
        # ------------------------------------------------------
        # Write the General information. 
        ws.merge_cells(f'A{NextRowIndex}:B{NextRowIndex}')
        cell = ws.cell(row=NextRowIndex, column=1, value='6th Degree Polynomial Model')
        cell.fill = Poly_fill
        cell.border = cell_border
        cell.font = Font(name="Arial", size=13, bold=True, color="000000")
        cell.alignment = center_alignment
        for i in range(1, 4):
            for j in range(2):
                cell1 = ws.cell(row=NextRowIndex+i, column=j+1, value='')
                cell1.fill = Poly_fill
        # Put the image. 
        Image_Obj = Read_Resize_Image('./assets/Polynomial Equation.png', 44)
        ws.add_image(Image_Obj, f"A{NextRowIndex+1}")
        NextRowIndex += 4
        # Write the 2PP model parameters. 
        for i, (header, value) in enumerate(zip(Labels_Poly, Values_Poly), start=NextRowIndex):
            # Write the title.
            cell1 = ws.cell(row=i, column=1, value=header + ':')
            cell1.fill = Poly_fill
            cell1.border = cell_border
            cell1.font = header_font
            cell1.alignment = left_alignment
            # Write the value. 
            cell1 = ws.cell(row=i, column=2, value=value)
            cell1.fill = Poly_fill
            cell1.border = cell_border
            cell1.font = cell_font
            cell1.alignment = left_alignment
        NextRowIndex = i + 2
        # ----------------------------------------------------------------------------------------------------------------------
        # Adding the raw data to the Excel file. 
        # Extract the data. 
        self.cursor.execute('SELECT Data, Data_shape, Data_dtype FROM HWTT WHERE id = ?', (ID,))
        Cont = self.cursor.fetchone()
        Data = Binary_to_Array(Cont[0], Cont[1], Cont[2])
        # Write the title. 
        ws.merge_cells(f'E1:G1')
        cell = ws.cell(row=1, column=5, value='Raw Data')
        cell.fill = Data_Fill
        cell.border = cell_border
        cell.font = Font(name="Arial", size=13, bold=True, color="000000")
        cell.alignment = center_alignment
        # Write the headers. 
        for j, header in enumerate(['Passes', 'Rut depth (mm)', 'Temperature (°C)'], start=5):
            cell = ws.cell(row=2, column=j, value=header)
            cell.fill = Data_Fill
            cell.border = cell_border
            cell.font = Font(name="Arial", size=12, bold=True, color="000000")
            cell.alignment = center_alignment
        # Write the raw data.
        ValidIndex = [Values_Info[16], Values_Info[17]]
        for i in range(Data.shape[1]):
            for j in range(3):
                cell = ws.cell(row=3 + i, column=5 + j, value=Data[j, i])
                if Data[0, i] < ValidIndex[0] or Data[0, i] > ValidIndex[1]:
                    cell.fill = Invalid_Fill
                else:
                    cell.fill = Data_Fill
                cell.border = cell_border
                cell.font = cell_font
                cell.alignment = center_alignment
                if j == 0:
                    cell.number_format = "0"
                else:
                    cell.number_format = "0.00"
        # --------------------------------------------------------------------------------------------------------------
        # Adjust the column dimensions. 
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["E"].width = 11
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 20
        # Save the Excel file. 
        wb.save(os.path.join(Directory, FileName))
        self.Export_ProgressBar.hide()
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Button_Export_Database(self):
        """
        This function creates a new Excel file, acts as a copy of all database. 
        """
        # Ask user for saving directory. 
        Directory = QFileDialog.getExistingDirectory(self, "Please select Saving Directory", "")
        # If a file is selected by the user, update the Input_SavePath.
        if not Directory:
            QMessageBox.critical(self, "Directory Selection Failed!", f"Directory was NOT selected. Please try again.")
            return
        print(f'Saving Directory: {Directory}')
        # # Freeze the main window. 
        # self.setEnabled(False)
        # Ask for the file name. 
        FileName, IsOkButtonPressed = QInputDialog.getText(self,
                                                           "Output File Name", 
                                                           "Please enter the output file name (without .xlsx):",
                                                           text=self.DB_Name)
        if IsOkButtonPressed:
            FileName = FileName + '.xlsx'
            print(f"Saving File Name: {FileName}")
        else:
            QMessageBox.critical(self, "Output File Name Failed!", 
                                 f"Output file name was NOT confirmed. Please try again.")
            return
        # --------------------------------------------------------------------------------------------------------------
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        # Set the title of the sheet
        ws.title = "Sheet1"
        # Define some styles. 
        Data_Fill       = PatternFill(start_color="A4FED3", end_color="A4FED3", fill_type="solid")
        Invalid_Fill    = PatternFill(start_color="FF1515", end_color="FF1515", fill_type="solid")
        thin            = Side(border_style="thin", color="000000")
        thick           = Side(border_style="thick", color="000000")
        cell_border     = Border(top=thin, left=thin, right=thin, bottom=thin)
        header_border   = Border(top=thick, left=thick, right=thick, bottom=thick)
        header_font     = Font(name="Arial", bold=True, size=11, color="000000")
        cell_font       = Font(name="Arial", size=11, color="000000")
        center_alignment= Alignment(horizontal="center", vertical="center")
        left_alignment  = Alignment(horizontal="left", vertical="center")
        # ----------------------------------------------
        # Write the general information. 
        ColNumbers = [1, 21, 43, 68]
        Titles     = ['General Information', 'Two-Part Power (2PP) model', 
                      'Yin et al. model', '6th Degree Polynomial model']
        for i, Range in enumerate(['A1:T1', 'U1:AP1', 'AQ1:BO1', 'BP1:CB1']):
            ws.merge_cells(Range)
            cell = ws.cell(row=1, column=ColNumbers[i], value=Titles[i])
            # cell.fill = Data_Fill
            cell.border = header_border
            cell.font = Font(name="Arial", size=16, bold=True, color="000000")
            cell.alignment = center_alignment
        RowIndex = 2
        # ----------------------------------------------
        # Extract the data. 
        ColNames = [
            'Bnumber', 'Lane_Num', 'Lift_Location', 'Lab_Aging', 'RepNumber', 'Wheel_Side', 'FileName', 'FileDirectory', 
            'Test_Name', 'Technician_Name', 'Test_Date', 'Test_Time', 'Test_Condition', 
            'Target_Test_Temperature', 'Avg_Test_Temperature', 'Std_Test_Temperature', 
            'Valid_Min_Pass', 'Valid_Max_Pass', 'Other_Comments', 'IsOutlier', 
            'TPP_StrippingNumber', 'TPP_Max_Rut_mm', 'TPP_Max_Pass', 
            'TPP_RuttingAt10k_mm', 'TPP_RuttingAt20k_mm', 
            'TPP_ModelCoeff_a', 'TPP_ModelCoeff_b', 
            'TPP_ModelCoeff_alpha', 'TPP_ModelCoeff_beta', 'TPP_ModelCoeff_gamma', 'TPP_ModelCoeff_Phi', 
            'TPP_Stripping_Rutting_mm', 'TPP_SIP', 'TPP_SIP_Yvalue', 'TPP_SIP_Adj', 'TPP_SIP_Adj_Yvalue', 
            'TPP_CreepLine_Slope', 'TPP_CreepLine_Intercept', 'TPP_StrippingLine_Slope', 'TPP_StrippingLine_Intercept', 
            'TPP_StrippingLine_Slope_Adj', 'TPP_StrippingLine_Intercept_Adj', 
            'Yin_StrippingNumber', 'Yin_RuttingAt10k_mm', 'Yin_RuttingAt20k_mm', 
            'Yin_Parameter_LCSN', 'Yin_Parameter_LCST', 'Yin_Parameter_DeltaEpsAt10k',
            'Yin_ModelCoeff_Step1_ro', 'Yin_ModelCoeff_Step1_LCult', 'Yin_ModelCoeff_Step1_beta', 
            'Yin_ModelCoeff_Step2_RutMax', 'Yin_ModelCoeff_Step2_alpha', 'Yin_ModelCoeff_Step2_lambda', 
            'Yin_ModelCoeff_Step3_Eps0', 'Yin_ModelCoeff_Step3_theta', 
            'Yin_Stripping_Rutting_mm', 'Yin_SIP', 'Yin_SIP_Yvalue', 'Yin_SIP_Adj', 'Yin_SIP_Adj_Yvalue', 
            'Yin_CreepLine_Slope', 'Yin_CreepLine_Intercept', 'Yin_StrippingLine_Slope', 'Yin_StrippingLine_Intercept', 
            'Yin_StrippingLine_Slope_Adj', 'Yin_StrippingLine_Intercept_Adj',
            'Poly6_StrippingNumber', 'Poly6_RuttingAt10k_mm', 'Poly6_RuttingAt20k_mm', 
            'Poly6_ModelCoeff_a0', 'Poly6_ModelCoeff_a1', 'Poly6_ModelCoeff_a2', 'Poly6_ModelCoeff_a3', 
            'Poly6_ModelCoeff_a4', 'Poly6_ModelCoeff_a5', 'Poly6_ModelCoeff_a6', 
            'Poly6_Stripping_Rutting_mm', 'Poly6_CreepLine_Slope', 'Poly6_CreepLine_Intercept']
        Labels   = [
            'B-number', 'Lane number', 'Lift location', 'Lab aging level', 'Repetition number', 
            'Testing wheel side', 'Raw data file name', 'Raw data directory', 
            'Test name (by operator)', 'Technician name', 'Testing date', 'Testing time', 'Testing condition (wet/dry)', 
            'Target testing temperature', 'Avg. of testing temperature', 'Std. of testing temperature', 
            'Index of first valid pass', 'Index of last valid pass', 
            'Any other comments/notes', 'Is this test considered Outlier', 
            'Stripping number (SN)', 'Maximum rutting (mm)', 'Maximum passes', 
            'Rutting @ 10k passes (mm)', 'Rutting @ 20k passes (mm)', 
            'Model coefficient, a', 'Model coefficient, b', 
            'Model coefficient, α', 'Model coefficient, β', 'Model coefficient, γ', 'Model coefficient, Φ', 
            'Stripping rutting (mm)', 'SIP', 'Rutting @ SIP (mm)', 'Adjusted SIP (@ 12.5 mm rutting)', 'Rutting @ adjusted SIP', 
            'Creep line slope', 'Creep line intercept (mm)', 'Stripping line slope', 'Stripping line intercept (mm)', 
            'Adjusted stripping line slope', 'Adjusted stripping line intercept (mm)', 
            'Stripping number (SN)', 'Rutting @ 10k passes (mm)', 'Rutting @ 20k passes (mm)', 
            'Analysis Parameter, LCSN', 'Analysis Parameter, LCST', 'Analysis Parameter, Δεp @ 10k',
            'Model coefficient, ρ', 'Model coefficient, LCult', 'Model coefficient, β', 
            'Model coefficient, Rut∞', 'Model coefficient, α', 'Model coefficient, λ', 
            'Model coefficient, ε0', 'Model coefficient, θ', 
            'Stripping rutting (mm)', 'SIP', 'Rutting @ SIP (mm)', 'Adjusted SIP (@ 12.5 mm rutting)', 'Rutting @ adjusted SIP', 
            'Creep line slope', 'Creep line intercept (mm)', 'Stripping line slope', 'Stripping line intercept (mm)', 
            'Adjusted stripping line slope', 'Adjusted stripping line intercept (mm)', 
            'Stripping number (SN)', 'Rutting @ 10k passes (mm)', 'Rutting @ 20k passes (mm)', 
            'Model coefficient, a0', 'Model coefficient, a1', 'Model coefficient, a2', 
            'Model coefficient, a3', 'Model coefficient, a4', 'Model coefficient, a5', 'Model coefficient, a6',
            'Stripping rutting (mm)', 'Creep line slope', 'Creep line intercept (mm)']
        NumFormat = [
            '0', '0', None, None, '0', None, None, None, None, None, None, None, None, '0.00', '0.00', '0.00', 
            '0', '0', None, '0', 
            '0', '0.00', '0', '0.00', '0.00', '0.0000', '0.0000', '0.0000E+00', '0.0000', '0.0', '0.0000', 
            '0.00', '0.0', '0.00', '0.0', '0.00', '0.00E+00', '0.00', '0.00E+00', '0.00', '0.00E+00', '0.00', 
            '0', '0.00', '0.00', '0.0', '0.0', '0.0000E+00', '0.0000', '0.0000', '0.0000', '0.0000', '0.0000', '0.0000', 
            '0.0000E+00', '0.0000E+00', '0.00', '0.0', '0.00', '0.0', '0.00', 
            '0.00E+00', '0.00', '0.00E+00', '0.00', '0.00E+00', '0.00', '0', '0.00', '0.00', 
            '0.0000E+00', '0.0000E+00', '0.0000E+00', '0.0000E+00', '0.0000E+00', '0.0000E+00', '0.0000E+00',
            '0.00', '0.00E+00', '0.00']
        self.cursor.execute(f'SELECT {", ".join(ColNames)} FROM HWTT')
        Values = self.cursor.fetchall()
        # ----------------------------------------------
        # Write the headers. 
        for i, header in enumerate(Labels, start=1):
            cell = ws.cell(row=RowIndex, column=i, value=header)
            # cell.fill = Data_Fill
            cell.border = header_border
            cell.font = header_font
            cell.alignment = center_alignment
        RowIndex += 1
        # Write the data. 
        for j in range(len(Values)):
            IsOutlier = Values[j][19]
            for i, res in enumerate(Values[j], start=1):
                cell = ws.cell(row=RowIndex+j, column=i, value=res)
                if IsOutlier:
                    cell.fill = Invalid_Fill
                else:
                    cell.fill = Data_Fill
                cell.border = cell_border
                cell.font = cell_font
                cell.alignment = center_alignment
                if NumFormat[i - 1] != None:
                    cell.number_format = NumFormat[i - 1]
        # Set the column widths.
        Letters = [get_column_letter(i + 1) for i in range(len(ColNames))]
        Widths  = [
            11.00, 14.43, 12.86, 25.29, 20.14, 20.43, 29.71, 28.14, 35.57, 18.43, 
            13.86, 13.71, 28.57, 28.86, 29.14, 28.71, 24.86, 24.43, 28.29, 31.71, 
            24.00, 23.57, 18.43, 28.86, 28.86, 21.43, 21.43, 21.43, 21.14, 21.86, 
            23.43, 23.00, 10.00, 20.86, 34.71, 24.86, 18.00, 27.29, 21.00, 30.43, 
            30.57, 39.86, 24.00, 28.86, 28.86, 27.71, 27.57, 32.57, 21.43, 25.71, 
            21.43, 25.29, 21.43, 21.14, 22.29, 21.14, 23.43,  8.43, 20.86, 34.71, 
            24.86, 18.00, 27.29, 21.00, 30.43, 30.57, 39.86, 24.00, 28.86, 28.86, 
            22.43, 23.00, 23.00, 23.00, 23.00, 23.00, 23.00, 23.00, 23.43, 27.29]
        for i in range(len(ColNames)):
            ws.column_dimensions[Letters[i]].width = Widths[i]    
        # Save the Excel file. 
        wb.save(os.path.join(Directory, FileName))
    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------